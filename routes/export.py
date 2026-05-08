"""Blueprint: export, dashboard, stores, batch-review, anomalies, termination."""
import json as _json
import time
import urllib.parse
from datetime import datetime as _dt, timedelta as _td, timezone as _tz

from flask import Blueprint, request, jsonify, session, Response

from auth import login_required, require_module
from config import TW_TZ
from db import (
    get_db,
    _dashboard_cache, _DASHBOARD_TTL,
    _labor_cost_cache, _LABOR_TTL,
    _heatmap_cache,
    _anomalies_cache, _ANOMALIES_TTL,
    _stores_cache, _STATIC_TTL,
    _badges_cache,
)
from notifications import _notify_review_result

bp = Blueprint('export', __name__)


def init():
    pass


# ── EDI helpers ───────────────────────────────────────────────────────────────

def _get_insurance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM insurance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}


def _roc_date(date_str):
    if not date_str:
        return '0000000'
    try:
        from datetime import date as _d
        d = _d.fromisoformat(str(date_str)[:10])
        return f'{d.year - 1911:03d}{d.month:02d}{d.day:02d}'
    except Exception:
        return '0000000'


def _edi_bytes(val, width, numeric=False):
    s = str(val or '')
    if numeric:
        return s.rjust(width, '0').encode('ascii', errors='replace')[:width]
    try:
        b = s.encode('big5', errors='replace')
    except Exception:
        b = s.encode('ascii', errors='replace')
    if len(b) < width:
        b = b + b' ' * (width - len(b))
    return b[:width]


def _get_edi_staff(staff_ids_str):
    with get_db() as conn:
        if staff_ids_str:
            ids = [int(x) for x in staff_ids_str.split(',') if x.strip().isdigit()]
            rows = conn.execute(
                "SELECT * FROM punch_staff WHERE id = ANY(%s) AND active=TRUE ORDER BY name",
                (ids,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()
    return rows


# ── Excel export helpers ──────────────────────────────────────────────────────

def _excel_response(wb, filename):
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _make_wb_ws(title):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    hdr_fill = PatternFill('solid', fgColor='0F1C3A')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill = PatternFill('solid', fgColor='F4F6FA')
    center   = Alignment(horizontal='center', vertical='center')
    return wb, ws, hdr_fill, hdr_font, alt_fill, center


# ═══════════════════════════════════════════════════════════════════
# Excel Export Routes
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    from openpyxl.styles import Font, Alignment, PatternFill
    month    = request.args.get('month', '')
    staff_id = request.args.get('staff_id', '')
    if not month:
        from datetime import date as _de
        month = _de.today().strftime('%Y-%m')

    conds, params = ["TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s"], [month]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ps.employee_code, ps.name as staff_name, ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   pr.punch_type,
                   to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') as punch_time,
                   pr.is_manual, pr.manual_by, pr.gps_distance, pr.location_name, pr.note
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ps.name, pr.punched_at
        """, params).fetchall()

    PUNCH_LABEL = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}
    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"出勤明細_{month}")

    headers = ['員工代碼','姓名','部門','職稱','日期','打卡類型','時間','補打卡','操作人','GPS距離(m)','地點','備註']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
                str(r['work_date']), PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
                r['punch_time'], '是' if r['is_manual'] else '', r['manual_by'] or '',
                r['gps_distance'] if r['gps_distance'] is not None else '',
                r['location_name'] or '', r['note'] or '']
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')

    for col, w in enumerate([10,12,10,10,12,10,8,6,10,12,16,20], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, f'attendance_{month}.xlsx')


@bp.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    from openpyxl.styles import Alignment
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _df
        month = _df.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.employee_code, ps.name, ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                   BOOL_OR(pr.is_manual) as has_manual, COUNT(*) as punch_count
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"出勤摘要_{month}")
    headers = ['員工代碼','姓名','部門','職稱','日期','上班','下班','工時(h)','打卡次數','含補打']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            from datetime import datetime as _dtx
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dtx.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dtx.fromisoformat(str(r['co_ts']))
                dur_h = round((co - ci).total_seconds() / 3600, 2)
            except Exception:
                pass
        vals = [r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
                str(r['work_date']), r['clock_in'] or '', r['clock_out'] or '',
                dur_h, r['punch_count'], '是' if r['has_manual'] else '']
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')

    for col, w in enumerate([10,12,10,10,12,8,8,8,8,8], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, f'attendance_summary_{month}.xlsx')


@bp.route('/api/attendance/anomaly-report', methods=['GET'])
@login_required
def api_anomaly_report_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import calendar as _cal
    from datetime import date as _dax, timedelta as _tdax, datetime as _dtx

    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, ps.department, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (month,)).fetchall()

        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, st.start_time::text, st.end_time::text,
                   ps.name as staff_name, ps.department
            FROM shift_assignments sa JOIN shift_types st ON st.id=sa.shift_type_id
            JOIN punch_staff ps ON ps.id=sa.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(sa.shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

        days_in  = _cal.monthrange(y, mo)[1]
        first_day = f"{y}-{mo:02d}-01"
        last_day  = f"{y}-{mo:02d}-{days_in:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date FROM leave_requests
            WHERE status='approved' AND start_date<=%s AND end_date>=%s
        """, (last_day, first_day)).fetchall()

        punched_dates = conn.execute("""
            SELECT DISTINCT pr.staff_id, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date
            FROM punch_records pr
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
        """, (month,)).fetchall()

    shift_map   = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    punched_set = {(r['staff_id'], str(r['work_date'])) for r in punched_dates}
    leave_set   = set()
    for lr in leave_rows:
        cur = _dax.fromisoformat(str(lr['start_date']))
        end = _dax.fromisoformat(str(lr['end_date']))
        while cur <= end:
            leave_set.add((lr['staff_id'], str(cur)))
            cur += _tdax(days=1)

    today = _dax.today()
    anomalies = []

    for r in punch_rows:
        ds = str(r['work_date']); sid = r['staff_id']
        shift = shift_map.get((sid, ds))
        anomaly_type = ''; detail = ''

        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _dax.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t  = str(r['clock_in'])[11:16];  co_t  = str(r['clock_out'])[11:16]
            sh_s  = str(shift['start_time'])[:5]; sh_e  = str(shift['end_time'])[:5]
            try:
                ci_m  = int(ci_t[:2])*60+int(ci_t[3:5])
                sh_s_m = int(sh_s[:2])*60+int(sh_s[3:5])
                if ci_m - sh_s_m > 10:
                    anomaly_type = '遲到'; detail = f"應 {sh_s}，實際 {ci_t}（+{ci_m-sh_s_m}分）"
            except Exception:
                pass
            if not anomaly_type:
                try:
                    co_m  = int(co_t[:2])*60+int(co_t[3:5])
                    sh_e_m = int(sh_e[:2])*60+int(sh_e[3:5])
                    if sh_e_m - co_m > 15:
                        anomaly_type = '早退'; detail = f"應 {sh_e}，實際 {co_t}（-{sh_e_m-co_m}分）"
                except Exception:
                    pass

        if anomaly_type:
            anomalies.append({
                'staff_name': r['staff_name'], 'department': r['department'] or '',
                'date': ds,
                'shift_start': str(shift['start_time'])[:5] if shift else '—',
                'shift_end':   str(shift['end_time'])[:5]   if shift else '—',
                'clock_in':    str(r['clock_in'])[11:16]  if r['clock_in']  else '—',
                'clock_out':   str(r['clock_out'])[11:16] if r['clock_out'] else '—',
                'anomaly_type': anomaly_type, 'detail': detail,
            })

    for sr in shift_rows:
        ds = str(sr['shift_date']); sid = sr['staff_id']
        if (sid, ds) not in punched_set and (sid, ds) not in leave_set:
            if _dax.fromisoformat(ds) < today:
                anomalies.append({
                    'staff_name': sr['staff_name'], 'department': sr['department'] or '',
                    'date': ds,
                    'shift_start': str(sr['start_time'])[:5], 'shift_end': str(sr['end_time'])[:5],
                    'clock_in': '—', 'clock_out': '—',
                    'anomaly_type': '完全未打卡',
                    'detail': f"排班 {str(sr['start_time'])[:5]}～{str(sr['end_time'])[:5]}，當日無任何打卡記錄",
                })
    anomalies.sort(key=lambda x: (x['date'], x['staff_name']))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'{month} 異常明細'
    thin = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin',  color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'),
    )
    header_fill  = PatternFill('solid', fgColor='0F1C3A')
    warn_fill    = PatternFill('solid', fgColor='FFF3CD')
    err_fill     = PatternFill('solid', fgColor='FDECEA')
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ['員工姓名','部門','日期','應上班','應下班','實際上班','實際下班','異常類型','說明']
    col_w   = [12, 10, 12, 8, 8, 8, 8, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = header_fill; cell.alignment = center_align; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    for ri, a in enumerate(anomalies, 2):
        row_fill = err_fill if a['anomaly_type'] in ('缺上班打卡','缺下班打卡','完全未打卡') else warn_fill
        vals = [a['staff_name'], a['department'], a['date'],
                a['shift_start'], a['shift_end'], a['clock_in'], a['clock_out'],
                a['anomaly_type'], a['detail']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = row_fill; cell.border = thin
            cell.alignment = center_align if ci != 9 else Alignment(vertical='center')

    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('摘要'); ws2.append(['統計', '數量']); ws2.append(['異常總筆數', len(anomalies)])
    by_type = {}
    for a in anomalies:
        by_type[a['anomaly_type']] = by_type.get(a['anomaly_type'], 0) + 1
    for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
        ws2.append([t, c])
    return _excel_response(wb, f'anomaly_{month}.xlsx')


@bp.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    from openpyxl.styles import Alignment
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dg
        month = _dg.today().strftime('%Y-%m')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code,
                   ps.department, ps.role, ps.salary_type
            FROM salary_records sr JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month=%s ORDER BY ps.name
        """, (month,)).fetchall()

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"薪資_{month}")
    headers = ['員工代碼','姓名','部門','職稱','薪資制度',
               '工作日','出勤天數','請假天數','無薪假天數',
               '津貼合計','扣除合計','加班費','實領金額','狀態','備註']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        sal_type = r['salary_type'] or 'monthly'
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
                '時薪制' if sal_type == 'hourly' else '月薪制',
                float(r['work_days'] or 0), float(r['actual_days'] or 0),
                float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
                float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
                float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
                '已確認' if r['status'] == 'confirmed' else '草稿', r['note'] or '']
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([10,12,10,10,8,8,8,8,8,10,10,10,12,8,20], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, f'salary_{month}.xlsx')


@bp.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    from openpyxl.styles import Alignment
    month    = request.args.get('month', '')
    year     = request.args.get('year', '')
    staff_id = request.args.get('staff_id', '')

    conds, params = ['lr.status=%s'], ['approved']
    if month:    conds.append("to_char(lr.start_date,'YYYY-MM')=%s"); params.append(month)
    if year:     conds.append("EXTRACT(YEAR FROM lr.start_date)=%s"); params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s"); params.append(int(staff_id))

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr JOIN punch_staff ps ON ps.id=lr.staff_id
            JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE {' AND '.join(conds)}
            ORDER BY lr.start_date, ps.name
        """, params).fetchall()

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"請假_{month or year or 'all'}")
    headers = ['員工代碼','姓名','部門','假別','薪資倍率','開始日期','結束日期','天數','原因','代理人','狀態']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    PAY_LABEL    = {1.0:'全薪', 0.5:'半薪', 0.0:'無薪'}
    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}
    for i, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '',
                r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
                str(r['start_date']), str(r['end_date']), float(r['total_days']),
                r['reason'] or '', r['substitute_name'] or '',
                STATUS_LABEL.get(r['status'], r['status'])]
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([10,12,10,10,8,12,12,6,24,10,8], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    tag = month or year or 'all'
    return _excel_response(wb, f'leave_{tag}.xlsx')


@bp.route('/api/export/monthly-stats', methods=['GET'])
@login_required
def api_export_monthly_stats():
    from openpyxl.styles import Alignment
    from datetime import datetime as _dtx
    month = request.args.get('month', '')
    if not month:
        from datetime import date as _dm
        month = _dm.today().strftime('%Y-%m')

    with get_db() as conn:
        detail = conn.execute("""
            SELECT ps.name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.name, ps.department, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()

    stats = {}
    for r in detail:
        key = (r['name'], r['department'] or '')
        if key not in stats:
            stats[key] = {'days': 0, 'total_min': 0}
        stats[key]['days'] += 1
        if r['ci'] and r['co']:
            try:
                ci = r['ci'] if hasattr(r['ci'], 'timestamp') else _dtx.fromisoformat(str(r['ci']))
                co = r['co'] if hasattr(r['co'], 'timestamp') else _dtx.fromisoformat(str(r['co']))
                stats[key]['total_min'] += int((co - ci).total_seconds() / 60)
            except Exception:
                pass

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"出勤月統計_{month}")
    headers = ['姓名','部門','出勤天數','總工時(h)','平均時數/天']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, (key, s) in enumerate(sorted(stats.items()), 2):
        name, dept = key
        days    = s['days']
        total_h = round(s['total_min'] / 60, 2)
        avg_h   = round(total_h / days, 2) if days else 0
        for col, v in enumerate([name, dept, days, total_h, avg_h], 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([14,12,10,10,12], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, f'monthly_stats_{month}.xlsx')


@bp.route('/api/export/overtime', methods=['GET'])
@login_required
def api_export_overtime():
    from openpyxl.styles import Alignment
    month  = request.args.get('month', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:  conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    if status: conds.append("r.status=%s"); params.append(status)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role as staff_role
            FROM overtime_requests r JOIN punch_staff ps ON ps.id=r.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY r.request_date DESC, r.created_at DESC
        """, params).fetchall()

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws(f"加班申請_{month or 'all'}")
    headers = ['員工代碼','姓名','部門','職稱','申請日期','開始時間','結束時間','加班時數','加班費','日別','原因','狀態','審核人','備註']
    DAY_LABEL    = {'weekday':'平日','holiday':'假日','rest_day':'休息日','sunday':'星期日'}
    STATUS_LABEL = {'approved':'已核准','rejected':'已退回','pending':'待審核'}
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        vals = [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['staff_role'] or '',
                str(r['request_date']),
                str(r['start_time'])[:5] if r['start_time'] else '',
                str(r['end_time'])[:5]   if r['end_time']   else '',
                float(r['ot_hours'] or 0), float(r['ot_pay'] or 0),
                DAY_LABEL.get(r['day_type'] or '', r['day_type'] or ''),
                r['reason'] or '', STATUS_LABEL.get(r['status'], r['status']),
                r['reviewed_by'] or '', r['review_note'] or '']
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([10,12,10,10,12,8,8,8,10,8,20,8,10,20], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, f'overtime_{month or "all"}.xlsx')


@bp.route('/api/export/training', methods=['GET'])
@login_required
def api_export_training():
    from openpyxl.styles import Alignment, PatternFill
    from datetime import date as _td, datetime as _dtx2
    staff_id = request.args.get('staff_id', '')
    category = request.args.get('category', '')
    sql = """
        SELECT tr.*, ps.name AS staff_name, ps.department
        FROM training_records tr JOIN punch_staff ps ON tr.staff_id=ps.id
        WHERE 1=1
    """
    params = []
    if staff_id: sql += " AND tr.staff_id=%s"; params.append(int(staff_id))
    if category: sql += " AND tr.category=%s"; params.append(category)
    sql += " ORDER BY ps.name, tr.expiry_date ASC NULLS LAST"
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()

    today = _td.today()
    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws("教育訓練記錄")
    exp_fill  = PatternFill('solid', fgColor='FFF0F0')
    soon_fill = PatternFill('solid', fgColor='FFFBEA')
    headers = ['姓名','部門','課程名稱','類別','完成日期','到期日期','證書號碼','狀態','剩餘天數']
    CAT_LABEL    = {'food_safety':'食品安全','fire_safety':'消防安全','first_aid':'急救訓練',
                    'hygiene':'衛生管理','service':'服務禮儀','equipment':'設備操作',
                    'general':'一般訓練','other':'其他'}
    STATUS_LABEL = {'expired':'已過期','expiring_soon':'即將到期','valid':'有效','no_expiry':'無到期日'}
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        expiry_str = str(r['expiry_date']) if r['expiry_date'] else ''
        days_left = ''; status = 'no_expiry'
        row_fill = alt_fill if i % 2 == 0 else None
        if expiry_str:
            ed = _dtx2.strptime(expiry_str, '%Y-%m-%d').date()
            dl = (ed - today).days
            days_left = dl
            if dl < 0:    status = 'expired';       row_fill = exp_fill
            elif dl <= 60: status = 'expiring_soon'; row_fill = soon_fill
            else:          status = 'valid'
        vals = [r['staff_name'], r['department'] or '',
                r['course_name'], CAT_LABEL.get(r['category'] or '', r['category'] or ''),
                str(r['completed_date']) if r['completed_date'] else '',
                expiry_str, r['certificate_no'] or '',
                STATUS_LABEL.get(status, status), days_left]
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if row_fill: c.fill = row_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([12,10,24,10,12,12,14,10,8], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, 'training_records.xlsx')


@bp.route('/api/export/expense-claims', methods=['GET'])
@login_required
def api_export_expense_claims():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    status     = request.args.get('status', '')
    year_month = request.args.get('ym', '')
    conds, params = ['TRUE'], []
    if status:     conds.append("ec.status=%s"); params.append(status)
    if year_month: conds.append("TO_CHAR(ec.expense_date,'YYYY-MM')=%s"); params.append(year_month)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code, ps.department
            FROM expense_claims ec JOIN punch_staff ps ON ps.id=ec.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY ec.expense_date ASC, ec.created_at ASC
        """, params).fetchall()

    def roc_ym(d):
        if not d: return ''
        try:
            from datetime import date as _date
            if hasattr(d, 'year'): return f"{d.year - 1911}/{d.month:02d}月"
            parts = str(d)[:7].split('-')
            return f"{int(parts[0]) - 1911}/{parts[1]}月"
        except Exception:
            return str(d)[:7]

    STATUS_LABEL = {'approved': '已核准', 'rejected': '已拒絕', 'pending': '待審核'}
    navy = '0F1C3A'; white = 'FFFFFF'; alt_gray = 'F7F8FB'; border_c = 'CBD2E0'
    hdr_font  = Font(bold=True, color=white, size=10, name='微軟正黑體')
    body_font = Font(size=10, name='微軟正黑體')
    hdr_fill  = PatternFill('solid', fgColor=navy)
    alt_fill  = PatternFill('solid', fgColor=alt_gray)
    thin      = Side(style='thin', color=border_c)
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right',  vertical='center')

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '費用申請匯出'
    ws.merge_cells('A1:M1')
    title_c = ws.cell(1, 1, '進光設計　費用申請表')
    title_c.font = Font(bold=True, size=14, name='微軟正黑體', color=navy)
    title_c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = [('年月',8),('員工代碼',10),('申請人',12),('費用名目',24),('費用類別',14),
               ('費用性質',9),('報帳方式',9),('費用金額',12),('戶名',14),('銀行',14),
               ('帳號',18),('備註',20),('狀態',9)]
    for col, (h, _) in enumerate(headers, 1):
        c = ws.cell(2, col, h); c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center_align; c.border = full_border
    ws.row_dimensions[2].height = 22; ws.freeze_panes = 'A3'

    total_amount = 0
    for i, r in enumerate(rows, 3):
        amt = float(r['amount'] or 0); total_amount += amt
        fill = alt_fill if i % 2 == 0 else None
        row_vals = [
            (roc_ym(r['expense_date']), center_align), (r['employee_code'] or '', center_align),
            (r['staff_name'] or '', left_align), (r['title'] or '', left_align),
            (r['category'] or '', center_align), (r.get('expense_type') or '支出', center_align),
            (r.get('reimbursement_method') or '匯款', center_align), (amt, right_align),
            (r.get('account_holder') or '', left_align), (r.get('bank_name') or '', left_align),
            (r.get('bank_account') or '', left_align), (r['note'] or '', left_align),
            (STATUS_LABEL.get(r['status'], r['status']), center_align),
        ]
        for col, (val, align) in enumerate(row_vals, 1):
            c = ws.cell(i, col, val); c.font = body_font; c.alignment = align
            c.border = full_border
            if fill: c.fill = fill
            if col == 8 and isinstance(val, (int, float)): c.number_format = '$#,##0'
        ws.row_dimensions[i].height = 18

    total_row = len(rows) + 3
    ws.merge_cells(f'A{total_row}:G{total_row}')
    tc = ws.cell(total_row, 1, '合　計')
    tc.font = Font(bold=True, size=10, name='微軟正黑體', color=navy)
    tc.alignment = center_align; tc.fill = PatternFill('solid', fgColor='E8ECF6'); tc.border = full_border
    for col in range(2, 8):
        ws.cell(total_row, col).border = full_border
        ws.cell(total_row, col).fill   = PatternFill('solid', fgColor='E8ECF6')
    ac = ws.cell(total_row, 8, total_amount)
    ac.font = Font(bold=True, size=10, name='微軟正黑體'); ac.alignment = right_align
    ac.number_format = '$#,##0'; ac.fill = PatternFill('solid', fgColor='E8ECF6'); ac.border = full_border
    for col in range(9, 14):
        ws.cell(total_row, col).border = full_border
        ws.cell(total_row, col).fill   = PatternFill('solid', fgColor='E8ECF6')
    ws.row_dimensions[total_row].height = 22

    for col, (_, w) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1; ws.print_title_rows = '1:2'

    fname = f"進光設計_費用申請_{year_month or status or 'all'}.xlsx"
    from io import BytesIO
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{urllib.parse.quote(fname)}"})


@bp.route('/api/export/performance', methods=['GET'])
@login_required
def api_export_performance():
    from openpyxl.styles import Alignment
    staff_id = request.args.get('staff_id', '')
    period   = request.args.get('period', '')
    conds, params = ['TRUE'], []
    if staff_id: conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    if period:   conds.append("pr.period_label ILIKE %s"); params.append(f'%{period}%')

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT pr.*, ps.name AS staff_name, ps.role AS staff_role, ps.department,
                   pt.name AS tpl_name
            FROM performance_reviews pr JOIN punch_staff ps ON ps.id=pr.staff_id
            LEFT JOIN performance_templates pt ON pt.id=pr.template_id
            WHERE {' AND '.join(conds)}
            ORDER BY pr.reviewed_at DESC
        """, params).fetchall()

    wb, ws, hdr_fill, hdr_font, alt_fill, center = _make_wb_ws("績效考核")
    headers = ['姓名','部門','職稱','考核期間','考核範本','分數','等級','備註','考核人','考核日期','薪資調整']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        reviewed_at = str(r['reviewed_at'])[:10] if r['reviewed_at'] else ''
        adj = r['salary_adjusted']
        adj_str = (f"+{adj:,.0f}" if adj and adj > 0 else (f"{adj:,.0f}" if adj and adj < 0 else '—'))
        vals = [r['staff_name'], r['department'] or '', r['staff_role'] or '',
                r['period_label'] or '', r['tpl_name'] or '',
                float(r['total_score'] or 0), r['grade'] or '',
                r['comments'] or '', r['reviewer'] or '', reviewed_at, adj_str]
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            if i % 2 == 0: c.fill = alt_fill
            c.alignment = Alignment(vertical='center')
    for col, w in enumerate([12,10,10,12,14,8,6,24,10,12,10], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    return _excel_response(wb, 'performance_reviews.xlsx')


# ═══════════════════════════════════════════════════════════════════
# Punch Request Review (single + batch)
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/punch/requests/<int:rid>', methods=['PUT'])
@login_required
def api_punch_req_review(rid):
    from routes.punch import punch_req_row
    b           = request.get_json(force=True)
    action      = b.get('action')
    reviewed_by = b.get('reviewed_by', '').strip()
    review_note = b.get('review_note', '').strip()
    if action not in ('approve', 'reject'):
        return jsonify({'error': 'invalid action'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_requests SET status=%s, reviewed_by=%s, review_note=%s, reviewed_at=NOW()
            WHERE id=%s
            RETURNING *, (SELECT name FROM punch_staff WHERE id=staff_id) as staff_name
        """, (new_status, reviewed_by, review_note, rid)).fetchone()
        if not row: return ('', 404)
        if action == 'approve':
            conn.execute("""
                INSERT INTO punch_records (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                VALUES (%s,%s,%s,%s,TRUE,%s)
            """, (row['staff_id'], row['punch_type'], row['requested_at'],
                  f'補打卡申請 #{rid}：{row["reason"]}', reviewed_by))
    LABEL  = {'in':'上班打卡','out':'下班打卡','break_out':'休息開始','break_in':'休息結束'}
    dt_str = row['requested_at'].isoformat()[:16].replace('T', ' ')
    extra  = f"{LABEL.get(row['punch_type'],'')} {dt_str}"
    if review_note: extra += f"\n審核意見：{review_note}"
    _notify_review_result(row['staff_id'], '補打卡申請', action, extra)
    _badges_cache.clear()
    return jsonify(punch_req_row(row))


# ═══════════════════════════════════════════════════════════════════
# Dashboard APIs
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/dashboard', methods=['GET'])
@login_required
def api_dashboard():
    import calendar as _cal
    from datetime import date as _dd, datetime as _ddt, timezone as _tz2, timedelta as _tdd
    TW    = _tz2(_tdd(hours=8))
    today = _ddt.now(TW).date()

    req_month = request.args.get('month', '').strip()
    if req_month and len(req_month) == 7:
        month = req_month
    else:
        month = today.strftime('%Y-%m')

    _cache_key = f"{month}:{today.isoformat()}"
    now = time.time()
    cached = _dashboard_cache.get(_cache_key)
    if cached and now - cached['at'] < _DASHBOARD_TTL:
        return jsonify(cached['data'])

    from datetime import datetime as _ddt2
    _today_start = _ddt2(today.year, today.month, today.day, tzinfo=TW)
    _today_end   = _today_start + _tdd(days=1)

    with get_db() as conn:
        today_detail_rows = conn.execute("""
            SELECT ps.id, ps.name, ps.role,
                   MAX(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   COUNT(pr.id) as punch_count,
                   MAX(lt.name) as leave_name
            FROM punch_staff ps
            LEFT JOIN punch_records pr ON pr.staff_id=ps.id AND pr.punched_at>=%s AND pr.punched_at<%s
            LEFT JOIN leave_requests lr ON lr.staff_id=ps.id AND lr.status='approved'
              AND lr.start_date<=%s AND lr.end_date>=%s
            LEFT JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE ps.active=TRUE
            GROUP BY ps.id, ps.name, ps.role
            ORDER BY ps.name
        """, (_today_start, _today_end, today, today)).fetchall()

        total_staff    = len(today_detail_rows)
        clocked_in     = sum(1 for r in today_detail_rows if r['clock_in'])
        clocked_out    = sum(1 for r in today_detail_rows if r['clock_out'])
        on_leave_today = sum(1 for r in today_detail_rows if r['leave_name'] and not r['clock_in'])

        today_detail = []
        for r in today_detail_rows:
            if r['clock_in']:
                status, label = ('done','已下班') if r['clock_out'] else ('working','上班中')
            elif r['leave_name']:
                status, label = 'leave', r['leave_name']
            else:
                status, label = 'absent', '未出勤'
            today_detail.append({
                'id': r['id'], 'name': r['name'], 'role': r['role'] or '',
                'clock_in': r['clock_in'] or '', 'clock_out': r['clock_out'] or '',
                'punch_count': r['punch_count'], 'status': status, 'status_label': label,
            })

        _prow = conn.execute("""
            SELECT
                (SELECT COUNT(*) FROM punch_requests    WHERE status='pending') AS punch_cnt,
                (SELECT COUNT(*) FROM overtime_requests WHERE status='pending') AS ot_cnt,
                (SELECT COUNT(*) FROM schedule_requests WHERE status IN ('pending','modified_pending')) AS sched_cnt,
                (SELECT COUNT(*) FROM leave_requests    WHERE status='pending') AS leave_cnt
        """).fetchone()

        sal_rows = conn.execute("""
            SELECT COUNT(*) as total_count,
                   COUNT(*) FILTER (WHERE status='confirmed') as confirmed_count,
                   COALESCE(SUM(net_pay),0) as total_net,
                   COALESCE(SUM(allowance_total),0) as total_allow,
                   COALESCE(SUM(deduction_total),0) as total_deduct
            FROM salary_records WHERE month=%s
        """, (month,)).fetchone()

        days_in_month = _cal.monthrange(today.year, today.month)[1]
        daily_rows = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d,
                   COUNT(DISTINCT staff_id) as cnt
            FROM punch_records WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY (punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY d
        """, (month,)).fetchall()
        daily_map = {str(r['d']): r['cnt'] for r in daily_rows}
        daily_attendance = []
        for day in range(1, days_in_month + 1):
            ds = f"{month}-{day:02d}"
            dt = _dd(today.year, today.month, day)
            daily_attendance.append({'date':ds,'day':day,'count':daily_map.get(ds,0),
                                     'is_past':dt<=today,'weekday':dt.weekday()})

        leave_dist_rows = conn.execute("""
            SELECT lt.name, lt.color, COUNT(*) as cnt, COALESCE(SUM(lr.total_days),0) as days
            FROM leave_requests lr JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.status='approved' AND to_char(lr.start_date,'YYYY-MM')=%s
            GROUP BY lt.name, lt.color ORDER BY days DESC
        """, (month,)).fetchall()
        leave_distribution = [{'name':r['name'],'color':r['color'],'count':r['cnt'],'days':float(r['days'])}
                               for r in leave_dist_rows]

        ot_rank_rows = conn.execute("""
            SELECT ps.name, ps.role,
                   COALESCE(SUM(r.ot_pay),0) as total_pay, COALESCE(SUM(r.ot_hours),0) as total_hours
            FROM overtime_requests r JOIN punch_staff ps ON ps.id=r.staff_id
            WHERE r.status='approved' AND to_char(r.request_date,'YYYY-MM')=%s
            GROUP BY ps.name, ps.role ORDER BY total_pay DESC LIMIT 8
        """, (month,)).fetchall()
        ot_ranking = [{'name':r['name'],'role':r['role'] or '','pay':float(r['total_pay']),'hours':float(r['total_hours'])}
                      for r in ot_rank_rows]

    from datetime import date as _ddc
    cur_month = _ddc.today().strftime('%Y-%m')
    result = {
        'month': month, 'today': str(today), 'is_current_month': month == cur_month,
        'today_summary': {'total':total_staff,'working':clocked_in-clocked_out,
                          'clocked_in':clocked_in,'clocked_out':clocked_out,
                          'on_leave':on_leave_today,'absent':total_staff-clocked_in-on_leave_today},
        'today_detail': today_detail,
        'pending': {'punch':_prow['punch_cnt'],'ot':_prow['ot_cnt'],'sched':_prow['sched_cnt'],
                    'leave':_prow['leave_cnt'],'total':_prow['punch_cnt']+_prow['ot_cnt']+_prow['sched_cnt']+_prow['leave_cnt']},
        'salary_summary': {'total_count':sal_rows['total_count'],'confirmed_count':sal_rows['confirmed_count'],
                           'total_net':float(sal_rows['total_net']),'total_allow':float(sal_rows['total_allow']),
                           'total_deduct':float(sal_rows['total_deduct'])},
        'daily_attendance': daily_attendance,
        'leave_distribution': leave_distribution,
        'ot_ranking': ot_ranking,
    }
    _dashboard_cache[_cache_key] = {'data': result, 'at': now}
    return jsonify(result)


@bp.route('/api/dashboard/labor-cost', methods=['GET'])
@login_required
def api_dashboard_labor_cost():
    now = time.time()
    if _labor_cost_cache.get('data') is not None and now - _labor_cost_cache['at'] < _LABOR_TTL:
        return jsonify(_labor_cost_cache['data'])
    from datetime import date as _dlc
    today = _dlc.today()
    months = []
    for i in range(11, -1, -1):
        m = today.month - i; y = today.year
        while m <= 0: m += 12; y -= 1
        months.append(f'{y}-{m:02d}')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT month, COALESCE(SUM(net_pay),0) as total
            FROM salary_records WHERE month=ANY(%s)
            GROUP BY month
        """, (months,)).fetchall()
    cost_map = {r['month']: float(r['total']) for r in rows}
    result = {'months': months, 'labor_cost': [cost_map.get(m, 0) for m in months]}
    _labor_cost_cache['data'] = result; _labor_cost_cache['at'] = now
    return jsonify(result)


@bp.route('/api/dashboard/attendance-heatmap', methods=['GET'])
@login_required
def api_dashboard_attendance_heatmap():
    from datetime import date as _dah, timedelta as _tdah
    import calendar as _calh
    month = request.args.get('month', '') or _dah.today().strftime('%Y-%m')
    now = time.time()
    _hc = _heatmap_cache.get(month)
    if _hc and now - _hc['at'] < _DASHBOARD_TTL:
        return jsonify(_hc['data'])
    y, mo = int(month[:4]), int(month[5:7])
    days_in = _calh.monthrange(y, mo)[1]

    with get_db() as conn:
        total_staff = conn.execute("SELECT COUNT(*) as c FROM punch_staff WHERE active=TRUE").fetchone()['c']
        punch_rows  = conn.execute("""
            SELECT (punched_at AT TIME ZONE 'Asia/Taipei')::date as d, COUNT(DISTINCT staff_id) as cnt
            FROM punch_records WHERE punch_type='in'
              AND to_char(punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY d
        """, (month,)).fetchall()
        leave_rows = conn.execute("""
            SELECT lr.start_date, lr.end_date, COUNT(*) as cnt
            FROM leave_requests lr
            WHERE lr.status='approved'
              AND (TO_CHAR(lr.start_date,'YYYY-MM')=%s OR TO_CHAR(lr.end_date,'YYYY-MM')=%s)
            GROUP BY lr.start_date, lr.end_date
        """, (month, month)).fetchall()

    punch_map = {str(r['d']): int(r['cnt']) for r in punch_rows}
    leave_map = {}
    for lr in leave_rows:
        s = _dah.fromisoformat(str(lr['start_date'])); e = _dah.fromisoformat(str(lr['end_date']))
        cur = s
        while cur <= e:
            ds = str(cur)
            if ds.startswith(month): leave_map[ds] = leave_map.get(ds, 0) + 1
            cur += _tdah(days=1)

    days = []
    for d in range(1, days_in + 1):
        ds = f'{y}-{mo:02d}-{d:02d}'; cnt = punch_map.get(ds, 0)
        rate = round(cnt / total_staff, 3) if total_staff > 0 else 0
        days.append({'date':ds,'day_of_week':_dah(y,mo,d).weekday(),'count':cnt,
                     'attendance_rate':rate,'on_leave':leave_map.get(ds,0)})

    result = {'month': month, 'total_staff': total_staff, 'days': days}
    _heatmap_cache[month] = {'data': result, 'at': now}
    return jsonify(result)


@bp.route('/api/dashboard/leave-distribution', methods=['GET'])
@login_required
def api_dashboard_leave_distribution():
    from datetime import date as _dld
    year = request.args.get('year', str(_dld.today().year))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lt.name, lt.color, COUNT(*) as cnt, COALESCE(SUM(lr.total_days), 0) as days
            FROM leave_requests lr JOIN leave_types lt ON lt.id=lr.leave_type_id
            WHERE lr.status='approved' AND EXTRACT(YEAR FROM lr.start_date)=%s
            GROUP BY lt.name, lt.color ORDER BY days DESC
        """, (int(year),)).fetchall()
    total = sum(float(r['days']) for r in rows)
    return jsonify({'year': year, 'total_leave_days': total, 'breakdown': [
        {'name':r['name'],'color':r['color'] or '#4a7bda','days':float(r['days']),
         'count':int(r['cnt']),'pct':round(float(r['days'])/total*100,1) if total>0 else 0}
        for r in rows]})


# ═══════════════════════════════════════════════════════════════════
# Withholding / Insurance / EDI
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    from datetime import date as _dwh
    from routes.finance import _get_finance_settings
    year = request.args.get('year', str(_dwh.today().year))
    fmt  = request.args.get('format', 'html')

    fs = _get_finance_settings()
    company_name    = fs.get('company_name', '')
    company_tax_id  = fs.get('company_tax_id', '')
    company_address = fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total),0)     AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld),0) AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary),0)      AS avg_insured
            FROM salary_records sr JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address
            ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = [{'no':i,'name':r['name'],'national_id':r['national_id'] or '—','address':r['address'] or '—',
              'gross':float(r['gross_salary']),'supp_nhi':supp_nhi(r['gross_salary'],r['avg_insured']),
              'tax':float(r['tax_withheld'])}
             for i, r in enumerate(rows, 1)]

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號','姓名','身分證字號','地址','年度薪資合計','二代健保補充費','扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True,color='FFFFFF',size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center',vertical='center'); c.border = thin
        for col, w in zip('ABCDEFG', [5,12,14,30,16,16,12]):
            ws.column_dimensions[col].width = w
        for d in data:
            ws.append([d['no'],d['name'],d['national_id'],d['address'],d['gross'],d['supp_nhi'],d['tax']])
        return _excel_response(wb, f'withholding_{year}.xlsx')

    rows_html = ''.join(f"""
      <tr>
        <td style="text-align:center">{d['no']}</td><td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td>
      </tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head>
<meta charset="UTF-8"><title>{year}年度薪資扣繳憑單</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;font-weight:600;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
.note{{font-size:10px;color:#888;border-top:1px solid #ddd;padding-top:8px}}
@media print{{button{{display:none}}}}
</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單（所得類別 50）</h2>
<div class="meta">扣繳義務人：{company_name}　統一編號：{company_tax_id}　地址：{company_address}　製表日期：{_dwh.today().isoformat()}</div>
<table>
<thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th><th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="note">※ 本報表依薪資紀錄計算，二代健保補充費 = 超出投保薪資部分 × 2.11%。</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@bp.route('/api/insurance/settings', methods=['GET'])
@require_module('salary')
def api_insurance_settings_get():
    return jsonify(_get_insurance_settings())


@bp.route('/api/insurance/settings', methods=['PUT'])
@require_module('salary')
def api_insurance_settings_put():
    b = request.get_json(force=True)
    with get_db() as conn:
        for k in ('labor_insurance_no','health_insurance_no','employer_name','employer_id'):
            conn.execute(
                "INSERT INTO insurance_settings VALUES (%s,%s) ON CONFLICT (setting_key) DO UPDATE SET setting_value=EXCLUDED.setting_value",
                (k, str(b.get(k,'')).strip()))
    return jsonify({'ok': True})


@bp.route('/api/export/edi/labor-enroll', methods=['GET'])
@require_module('salary')
def api_edi_labor_enroll():
    event_type = request.args.get('event_type','in')
    staff_ids  = request.args.get('staff_ids','')
    event_date = request.args.get('event_date','')
    cfg        = _get_insurance_settings()
    labor_no   = cfg.get('labor_insurance_no','').ljust(8)[:8]
    event_code = b'1' if event_type=='in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M','男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6,'0').encode('ascii')
        lines.append(_edi_bytes(labor_no,8)+_edi_bytes(s['name'],20)+_edi_bytes(s.get('national_id',''),10)+
                     _roc_date(s.get('birth_date')).encode('ascii')+event_roc+event_code+insured+gender_code+b'00')
    fname = f'labor_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@bp.route('/api/export/edi/labor-salary', methods=['GET'])
@require_module('salary')
def api_edi_labor_salary():
    month     = request.args.get('month','')
    staff_ids = request.args.get('staff_ids','')
    cfg       = _get_insurance_settings()
    labor_no  = cfg.get('labor_insurance_no','').ljust(8)[:8]
    if not month:
        from datetime import date as _dm2
        month = _dm2.today().strftime('%Y-%m')
    month_roc = f"{int(month[:4])-1911:03d}{month[5:7]}".encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6,'0').encode('ascii')
        lines.append(_edi_bytes(labor_no,8)+_edi_bytes(s['name'],20)+_edi_bytes(s.get('national_id',''),10)+
                     insured+month_roc)
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename=labor_salary_{month}.edi'})


@bp.route('/api/export/edi/health-enroll', methods=['GET'])
@require_module('salary')
def api_edi_health_enroll():
    event_type = request.args.get('event_type','in')
    staff_ids  = request.args.get('staff_ids','')
    event_date = request.args.get('event_date','')
    cfg        = _get_insurance_settings()
    health_no  = cfg.get('health_insurance_no','').ljust(10)[:10]
    event_code = b'1' if event_type=='in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M','男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6,'0').encode('ascii')
        lines.append(_edi_bytes(health_no,10)+_edi_bytes(s['name'],20)+_edi_bytes(s.get('national_id',''),10)+
                     _roc_date(s.get('birth_date')).encode('ascii')+event_roc+event_code+insured+gender_code)
    fname = f'health_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


# ═══════════════════════════════════════════════════════════════════
# Stores
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/stores', methods=['GET'])
@login_required
def api_stores_list():
    now = time.time()
    cached = _stores_cache.get('data')
    if cached is not None and now - _stores_cache['at'] < _STATIC_TTL:
        return jsonify(cached)
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stores ORDER BY id").fetchall()
    result = [dict(r) for r in rows]
    _stores_cache['data'] = result; _stores_cache['at'] = now
    return jsonify(result)


@bp.route('/api/stores', methods=['POST'])
@login_required
def api_stores_create():
    b = request.get_json(force=True)
    name = (b.get('name') or '').strip()
    if not name: return jsonify({'error': '店名為必填'}), 400
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO stores (name, code, address) VALUES (%s,%s,%s) RETURNING *",
            (name, (b.get('code') or '').strip() or None, (b.get('address') or '').strip())
        ).fetchone()
    _stores_cache['data'] = None
    return jsonify(dict(row)), 201


@bp.route('/api/stores/<int:sid>', methods=['PUT'])
@login_required
def api_stores_update(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        row = conn.execute("""
            UPDATE stores SET name=%s, code=%s, address=%s, active=%s WHERE id=%s RETURNING *
        """, ((b.get('name') or '').strip(), b.get('code') or None,
              (b.get('address') or '').strip(), bool(b.get('active', True)), sid)).fetchone()
    _stores_cache['data'] = None
    return jsonify(dict(row)) if row else ('', 404)


@bp.route('/api/stores/<int:sid>', methods=['DELETE'])
@login_required
def api_stores_delete(sid):
    with get_db() as conn:
        conn.execute("UPDATE punch_staff     SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("UPDATE punch_locations SET store_id=NULL WHERE store_id=%s", (sid,))
        conn.execute("DELETE FROM stores WHERE id=%s", (sid,))
    _stores_cache['data'] = None
    return jsonify({'deleted': sid})


@bp.route('/api/stores/<int:sid>/staff', methods=['GET'])
@login_required
def api_store_staff(sid):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, role, active FROM punch_staff WHERE store_id=%s ORDER BY name", (sid,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/staff/<int:sid>/store', methods=['PUT'])
@login_required
def api_staff_assign_store(sid):
    b = request.get_json(force=True)
    with get_db() as conn:
        conn.execute("UPDATE punch_staff SET store_id=%s WHERE id=%s", (b.get('store_id'), sid))
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════
# Staffing Requirements & Auto-Schedule
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/shifts/staffing-requirements', methods=['GET'])
@login_required
def api_staffing_req_get():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.id, r.shift_type_id, r.day_of_week, r.required_count,
                   st.name as shift_name, st.color as shift_color
            FROM shift_staffing_requirements r JOIN shift_types st ON st.id=r.shift_type_id
            ORDER BY st.sort_order, r.day_of_week
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/shifts/staffing-requirements', methods=['PUT'])
@login_required
def api_staffing_req_put():
    items = request.get_json(force=True)
    if not isinstance(items, list): return jsonify({'error': '格式錯誤'}), 400
    count = 0
    with get_db() as conn:
        for it in items:
            stid = int(it.get('shift_type_id',0)); dow = int(it.get('day_of_week',0))
            req  = max(0, int(it.get('required_count',1)))
            if req == 0:
                conn.execute("DELETE FROM shift_staffing_requirements WHERE shift_type_id=%s AND day_of_week=%s", (stid,dow))
            else:
                conn.execute("""
                    INSERT INTO shift_staffing_requirements (shift_type_id, day_of_week, required_count, updated_at)
                    VALUES (%s,%s,%s,NOW())
                    ON CONFLICT (shift_type_id, day_of_week) DO UPDATE SET required_count=EXCLUDED.required_count, updated_at=NOW()
                """, (stid, dow, req))
            count += 1
    return jsonify({'ok': True, 'upserted': count})


@bp.route('/api/schedule/auto-generate', methods=['POST'])
@login_required
def api_auto_generate_schedule():
    from datetime import date as _dag, timedelta as _tdag
    import calendar as _calag

    b         = request.get_json(force=True)
    month     = (b.get('month') or '').strip() or _dag.today().strftime('%Y-%m')
    overwrite = bool(b.get('overwrite', False))
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    all_dates = [_dag(y, mo, d) for d in range(1, _calag.monthrange(y, mo)[1] + 1)]

    with get_db() as conn:
        shift_types  = conn.execute("SELECT * FROM shift_types WHERE active=TRUE ORDER BY sort_order").fetchall()
        requirements = conn.execute("SELECT shift_type_id, day_of_week, required_count FROM shift_staffing_requirements").fetchall()
        staff_list   = conn.execute("SELECT id, name FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()
        leave_rows   = conn.execute("""
            SELECT staff_id, start_date, end_date FROM leave_requests WHERE status='approved'
              AND start_date<=%s AND end_date>=%s
        """, (f'{y}-{mo:02d}-{_calag.monthrange(y,mo)[1]:02d}', f'{y}-{mo:02d}-01')).fetchall()
        sched_rows   = conn.execute("""
            SELECT staff_id, requested_dates FROM schedule_requests
            WHERE status='approved' AND to_char(created_at,'YYYY-MM')=%s
        """, (month,)).fetchall()
        existing = conn.execute("""
            SELECT staff_id, shift_date FROM shift_assignments WHERE TO_CHAR(shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()

    off_days = set()
    for lr in leave_rows:
        cur = _dag.fromisoformat(str(lr['start_date'])); end = _dag.fromisoformat(str(lr['end_date']))
        while cur <= end:
            off_days.add((lr['staff_id'], str(cur))); cur += _tdag(days=1)
    for sr in sched_rows:
        rdates = sr['requested_dates']
        if isinstance(rdates, str):
            try: rdates = _json.loads(rdates)
            except (ValueError, TypeError): rdates = []
        for ds in (rdates or []):
            off_days.add((sr['staff_id'], ds))

    existing_set  = {(r['staff_id'], str(r['shift_date'])) for r in existing}
    req_map       = {(r['shift_type_id'], r['day_of_week']): r['required_count'] for r in requirements}
    assigned_days = {s['id']: [] for s in staff_list}
    assignments   = []; conflicts = []
    staff_ids     = [s['id'] for s in staff_list]
    staff_name_map= {s['id']: s['name'] for s in staff_list}

    for date in all_dates:
        dow = date.weekday(); ds = str(date)
        for st in shift_types:
            stid   = st['id']; needed = req_map.get((stid, dow), 0)
            if needed <= 0: continue
            available = [sid for sid in staff_ids if (sid,ds) not in off_days]
            already_today = {a['staff_id'] for a in assignments if a['shift_date']==ds}
            available = [sid for sid in available if sid not in already_today]

            def consecutive_days(sid, d):
                days = sorted(assigned_days[sid]); streak = 0; check = d
                while check in days:
                    streak += 1; check = str(_dag.fromisoformat(check) - _tdag(days=1))
                return streak

            available_ok = sorted([sid for sid in available if consecutive_days(sid,ds)<6],
                                   key=lambda sid: len(assigned_days[sid]))
            assigned_count = 0
            for sid in available_ok:
                if assigned_count >= needed: break
                if not overwrite and (sid,ds) in existing_set:
                    assigned_count += 1; continue
                assignments.append({'staff_id':sid,'staff_name':staff_name_map[sid],
                                    'shift_type_id':stid,'shift_name':st['name'],'shift_date':ds})
                assigned_days[sid].append(ds); assigned_count += 1
            if assigned_count < needed:
                conflicts.append({'type':'understaffed','date':ds,'shift':st['name'],
                                  'detail':f'{ds} {st["name"]} 需要 {needed} 人，僅能排 {assigned_count} 人'})

    inserted = 0
    if assignments:
        with get_db() as conn:
            for a in assignments:
                try:
                    if overwrite:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s) ON CONFLICT (staff_id, shift_date)
                            DO UPDATE SET shift_type_id=EXCLUDED.shift_type_id
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    else:
                        conn.execute("""
                            INSERT INTO shift_assignments (staff_id, shift_type_id, shift_date)
                            VALUES (%s,%s,%s) ON CONFLICT DO NOTHING
                        """, (a['staff_id'], a['shift_type_id'], a['shift_date']))
                    inserted += 1
                except Exception:
                    pass

    return jsonify({'ok':True,'month':month,'assignments':assignments,'conflicts':conflicts,
                    'summary':{'assigned':inserted,'conflict_count':len(conflicts)}})


# ═══════════════════════════════════════════════════════════════════
# Salary PDF
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/salary/records/<int:rid>/pdf', methods=['GET'])
@require_module('salary')
def api_salary_pdf(rid):
    from routes.salary import salary_record_row
    if not session.get('logged_in'):
        sid = session.get('punch_staff_id')
        if not sid: return '未登入', 401
    with get_db() as conn:
        row = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code, ps.department,
                   ps.role, ps.salary_type, ps.hourly_rate, ps.hire_date
            FROM salary_records sr JOIN punch_staff ps ON ps.id=sr.staff_id
            WHERE sr.id=%s
        """, (rid,)).fetchone()
    if not row: return '找不到薪資記錄', 404
    if not session.get('logged_in') and row['staff_id'] != session.get('punch_staff_id'):
        return '無權限', 403

    d = salary_record_row(row)
    items        = d.get('items') or []
    allow_items  = [i for i in items if i.get('type') == 'allowance']
    deduct_items = [i for i in items if i.get('type') == 'deduction']
    is_hourly    = (row['salary_type'] == 'hourly')

    def money(v):
        try: return f"${float(v):,.0f}"
        except (ValueError, TypeError): return '$0'

    def esc_h(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

    allow_rows = ''.join(f"<tr><td>{esc_h(i['name'])}</td><td class='num green'>{money(i['amount'])}</td><td class='note'>{esc_h(i.get('calc_note',''))}</td></tr>"
                         for i in allow_items)
    deduct_rows = ''.join(f"<tr><td>{esc_h(i['name'])}</td><td class='num red'>-{money(i['amount'])}</td><td class='note'>{esc_h(i.get('calc_note',''))}</td></tr>"
                          for i in deduct_items)
    punch_table = ''
    if is_hourly and d.get('punch_details'):
        punch_rows_html = ''.join(f"<tr><td>{p['date']}</td><td>{p['clock_in']}</td><td>{p['clock_out']}</td><td>{p.get('break_mins',0)} min</td><td class='num'>{p['net_hours']} h</td></tr>"
                                  for p in d['punch_details'])
        punch_table = f"<h3>每日工時明細</h3><table><thead><tr><th>日期</th><th>上班</th><th>下班</th><th>休息</th><th>工時</th></tr></thead><tbody>{punch_rows_html}</tbody><tfoot><tr><td colspan='4'><strong>合計</strong></td><td class='num'><strong>{d.get('actual_work_hours',0)} h</strong></td></tr></tfoot></table>"

    status_str = '已確認' if row['status'] == 'confirmed' else '草稿（未確認）'
    sal_type   = '時薪制' if is_hourly else '月薪制'
    attend_str = (f"實際工時 {d.get('actual_work_hours',0)}h × 時薪 ${float(row['hourly_rate'] or 0):,.0f}"
                  if is_hourly else f"出勤 {d.get('actual_days',0)} 天 / 工作日 {d.get('work_days',0)} 天")
    if float(d.get('leave_days',0)) > 0: attend_str += f"，請假 {d.get('leave_days',0)} 天"
    if float(d.get('unpaid_days',0)) > 0: attend_str += f"（無薪 {d.get('unpaid_days',0)} 天）"

    html = f"""<!DOCTYPE html><html lang="zh-Hant"><head><meta charset="utf-8">
<title>薪資單 {esc_h(row['staff_name'])} {esc_h(row['month'])}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans TC','PingFang TC','Microsoft JhengHei',sans-serif;font-size:13px;color:#1a2340;background:#fff;padding:32px}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #1a2340;padding-bottom:16px;margin-bottom:24px}}
.company{{font-size:20px;font-weight:800;color:#1a2340}}.slip-title{{font-size:14px;color:#666;margin-top:4px}}
.staff-info{{font-size:12px;color:#444;text-align:right;line-height:1.8}}
.summary{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:24px}}
.sum-card{{border:1.5px solid #e2e8f0;border-radius:8px;padding:12px 16px;text-align:center}}
.sum-label{{font-size:10px;color:#888;margin-bottom:4px;text-transform:uppercase;letter-spacing:.06em}}
.sum-val{{font-size:22px;font-weight:800;font-family:'DM Mono',monospace}}
.sum-val.green{{color:#2e9e6b}}.sum-val.red{{color:#d64242}}.sum-val.navy{{color:#1a2340}}
.attend{{background:#f8fafc;border-radius:6px;padding:8px 14px;font-size:12px;color:#666;margin-bottom:20px}}
h3{{font-size:12px;font-weight:700;color:#888;letter-spacing:.08em;text-transform:uppercase;margin:20px 0 8px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#f1f5f9;padding:8px 12px;text-align:left;font-size:11px;font-weight:700;color:#666;border-bottom:2px solid #e2e8f0}}
td{{padding:7px 12px;border-bottom:1px solid #f0f2f8}}
td.num{{text-align:right;font-family:'DM Mono',monospace;font-weight:600}}
td.note{{font-size:11px;color:#999}}td.green{{color:#2e9e6b}}td.red{{color:#d64242}}
tfoot td{{font-weight:700;background:#f8fafc;border-top:2px solid #e2e8f0}}
.net-row td{{font-size:16px;font-weight:800;background:#1a2340;color:#fff}}
.net-row td.num{{color:#f0c040;font-size:20px}}
.footer{{margin-top:32px;padding-top:16px;border-top:1px solid #e2e8f0;display:flex;justify-content:space-between;font-size:11px;color:#999}}
.sign-area{{display:flex;gap:48px;margin-top:40px}}
.sign-box{{flex:1;border-top:1px solid #ccc;padding-top:6px;font-size:11px;color:#666}}
@media print{{body{{padding:16px}}@page{{margin:12mm;size:A4}}.no-print{{display:none !important}}}}
</style></head><body>
<div class="no-print" style="text-align:right;margin-bottom:20px">
  <button onclick="window.print()" style="padding:10px 24px;background:#1a2340;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer">列印 / 儲存 PDF</button>
</div>
<div class="header">
  <div><div class="company">薪資明細單</div><div class="slip-title">{esc_h(row['month'])} · {sal_type}</div></div>
  <div class="staff-info">
    <div><strong>{esc_h(row['staff_name'])}</strong></div>
    <div>{esc_h(row['employee_code'] or '')}　{esc_h(row['department'] or '')}　{esc_h(row['role'] or '')}</div>
    <div>到職日：{esc_h(str(row['hire_date']) if row['hire_date'] else '—')}</div>
    <div>狀態：<strong>{status_str}</strong></div>
  </div>
</div>
<div class="summary">
  <div class="sum-card"><div class="sum-label">津貼合計</div><div class="sum-val green">{money(d.get('allowance_total',0))}</div></div>
  <div class="sum-card"><div class="sum-label">扣除合計</div><div class="sum-val red">-{money(d.get('deduction_total',0))}</div></div>
  <div class="sum-card" style="border-color:#1a2340"><div class="sum-label">實領金額</div><div class="sum-val navy">{money(d.get('net_pay',0))}</div></div>
</div>
<div class="attend">{attend_str}</div>
<h3>津貼項目</h3>
<table><thead><tr><th>項目</th><th style="text-align:right">金額</th><th>計算說明</th></tr></thead>
<tbody>{allow_rows}</tbody>
<tfoot><tr><td><strong>津貼合計</strong></td><td class="num green"><strong>{money(d.get('allowance_total',0))}</strong></td><td></td></tr></tfoot></table>
<h3>扣除項目</h3>
<table><thead><tr><th>項目</th><th style="text-align:right">金額</th><th>計算說明</th></tr></thead>
<tbody>{deduct_rows if deduct_rows else '<tr><td colspan="3" style="color:#ccc;text-align:center;padding:12px">無扣除項目</td></tr>'}</tbody>
<tfoot><tr><td><strong>扣除合計</strong></td><td class="num red"><strong>-{money(d.get('deduction_total',0))}</strong></td><td></td></tr></tfoot></table>
<table style="margin-top:12px"><tbody>
  <tr class="net-row"><td><strong>實領金額</strong></td><td class="num">{money(d.get('net_pay',0))}</td>
  <td style="color:#ccc;font-size:11px">= 津貼 {money(d.get('allowance_total',0))} - 扣除 {money(d.get('deduction_total',0))}</td></tr>
</tbody></table>
{punch_table}
<div class="sign-area">
  <div class="sign-box">員工簽名</div><div class="sign-box">主管確認</div><div class="sign-box">人資確認</div>
</div>
<div class="footer"><span>本薪資單由系統自動產生</span>
<span>列印日期：<script>document.write(new Date().toLocaleDateString('zh-TW'))</script></span></div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ═══════════════════════════════════════════════════════════════════
# Batch Review
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/punch/requests/batch', methods=['POST'])
@login_required
def api_punch_req_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action'); by = b.get('reviewed_by', '管理員'); note = b.get('review_note', '')
    if not ids or action not in ('approve','reject'): return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE punch_requests SET status=%s, reviewed_by=%s, review_note=%s, reviewed_at=NOW()
                WHERE id=%s AND status='pending' RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    conn.execute("""
                        INSERT INTO punch_records (staff_id, punch_type, punched_at, note, is_manual, manual_by)
                        VALUES (%s,%s,%s,%s,TRUE,%s)
                    """, (row['staff_id'], row['punch_type'], row['requested_at'], f'補打卡申請#{rid}', by))
                _notify_review_result(row['staff_id'], '補打卡申請', action,
                                      f'批次審核意見：{note}' if note else '')
                done += 1
    return jsonify({'ok': True, 'done': done})


@bp.route('/api/overtime/requests/batch', methods=['POST'])
@login_required
def api_ot_batch():
    from routes.overtime import _calc_ot_pay
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action'); by = b.get('reviewed_by', '管理員'); note = b.get('review_note', '')
    if not ids or action not in ('approve','reject'): return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE overtime_requests SET status=%s, reviewed_by=%s, review_note=%s, reviewed_at=NOW()
                WHERE id=%s AND status='pending' RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                pay = 0
                if action == 'approve':
                    pay, _ = _calc_ot_pay(dict(row), float(row['ot_hours']), row.get('day_type','weekday'))
                    conn.execute("UPDATE overtime_requests SET ot_pay=%s WHERE id=%s", (pay, rid))
                time_str = (f"{row['start_time']}～{row['end_time']}"
                            if row.get('start_time') and row.get('end_time')
                            else f"{float(row['ot_hours'])} 小時")
                extra = f"{row['request_date']} {time_str}"
                if action == 'approve' and pay > 0: extra += f"\n加班費：${float(pay):,.0f}"
                if note: extra += f"\n審核意見：{note}"
                _notify_review_result(row['staff_id'], '加班申請', action, extra)
                done += 1
    _badges_cache.clear()
    return jsonify({'ok': True, 'done': done})


@bp.route('/api/schedule/requests/batch', methods=['POST'])
@login_required
def api_sched_batch():
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action'); by = b.get('reviewed_by', '管理員'); note = b.get('review_note', '')
    if not ids or action not in ('approve','reject'): return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        for rid in ids:
            row = conn.execute("""
                UPDATE schedule_requests SET status=%s, reviewed_by=%s, review_note=%s,
                  reviewed_at=NOW(), updated_at=NOW()
                WHERE id=%s AND status IN ('pending','modified_pending') RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                _notify_review_result(row['staff_id'], '排休申請', action, ''); done += 1
    return jsonify({'ok': True, 'done': done})


@bp.route('/api/leave/requests/batch', methods=['POST'])
@login_required
def api_leave_batch():
    from routes.leave import _update_leave_balance
    b      = request.get_json(force=True)
    ids    = [int(i) for i in b.get('ids', [])]
    action = b.get('action'); by = b.get('reviewed_by', '管理員'); note = b.get('review_note', '')
    if not ids or action not in ('approve','reject'): return jsonify({'error': '參數錯誤'}), 400
    new_status = 'approved' if action == 'approve' else 'rejected'
    done = 0
    with get_db() as conn:
        pending = {r['id']: r for r in conn.execute(
            "SELECT * FROM leave_requests WHERE id=ANY(%s) AND status='pending'", (ids,)
        ).fetchall()}
        for rid in ids:
            old = pending.get(rid)
            if not old: continue
            row = conn.execute("""
                UPDATE leave_requests SET status=%s, reviewed_by=%s, review_note=%s,
                  reviewed_at=NOW(), updated_at=NOW()
                WHERE id=%s RETURNING *
            """, (new_status, by, note, rid)).fetchone()
            if row:
                if action == 'approve':
                    _update_leave_balance(conn, old['staff_id'], old['leave_type_id'],
                                          str(old['start_date'])[:4], float(old['total_days']))
                extra = f"{str(old['start_date'])} ~ {str(old['end_date'])} 共 {float(old['total_days'])} 天"
                if note: extra += f"\n審核意見：{note}"
                _notify_review_result(old['staff_id'], '請假申請', action, extra); done += 1
    _badges_cache.clear()
    return jsonify({'ok': True, 'done': done})


# ═══════════════════════════════════════════════════════════════════
# Attendance Anomalies
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/attendance/anomalies', methods=['GET'])
@login_required
def api_attendance_anomalies():
    _an_now = time.time()
    if _anomalies_cache.get('data') is not None and _an_now - _anomalies_cache['at'] < _ANOMALIES_TTL:
        return jsonify(_anomalies_cache['data'])
    from datetime import date as _da, datetime as _dta, timezone as _tz2, timedelta as _td2
    TW    = _tz2(_td2(hours=8)); today = _dta.now(TW).date()
    date_from = today - _td2(days=7)

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name, ps.role, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   array_agg(pr.punch_type ORDER BY pr.punched_at) as types,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as first_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as last_out
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id
            WHERE (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date BETWEEN %s AND %s AND ps.active=TRUE
            GROUP BY ps.id, ps.name, ps.role, ps.department, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date DESC, ps.name
        """, (date_from, today)).fetchall()

        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, st.start_time, st.end_time, st.name as shift_name
            FROM shift_assignments sa JOIN shift_types st ON st.id=sa.shift_type_id
            WHERE sa.shift_date BETWEEN %s AND %s
        """, (date_from, today)).fetchall()
        shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}

        all_staff = conn.execute(
            "SELECT id, name, role, department FROM punch_staff WHERE active=TRUE"
        ).fetchall()
        today_punched_ids   = {r['staff_id'] for r in rows if str(r['work_date']) == str(today)}
        on_leave_today_ids  = set()
        for r in conn.execute("""
            SELECT DISTINCT staff_id FROM leave_requests
            WHERE status='approved' AND start_date<=%s AND end_date>=%s
        """, (today, today)).fetchall():
            on_leave_today_ids.add(r['staff_id'])

    anomalies = []
    for r in rows:
        types = list(r['types']) if r['types'] else []
        has_in = 'in' in types; has_out = 'out' in types; ds = str(r['work_date'])

        if has_in and not has_out and ds != str(today):
            anomalies.append({'type':'missing_out','label':'忘記下班打卡','severity':'warning',
                               'staff_id':r['staff_id'],'name':r['name'],'role':r['role'] or '',
                               'department':r['department'] or '','date':ds,
                               'detail':f"上班 {r['first_in']}，無下班記錄"})
        if not has_in and has_out:
            anomalies.append({'type':'missing_in','label':'忘記上班打卡','severity':'warning',
                               'staff_id':r['staff_id'],'name':r['name'],'role':r['role'] or '',
                               'department':r['department'] or '','date':ds,
                               'detail':f"下班 {r['last_out']}，無上班記錄"})
        if has_in and r['first_in']:
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['start_time']:
                try:
                    sh, sm = map(int, str(shift['start_time'])[:5].split(':'))
                    ih, im = map(int, r['first_in'].split(':'))
                    late = (ih*60+im)-(sh*60+sm)
                    if late > 10:
                        anomalies.append({'type':'late','label':'遲到','severity':'warning',
                                          'staff_id':r['staff_id'],'name':r['name'],'role':r['role'] or '',
                                          'department':r['department'] or '','date':ds,
                                          'detail':f"應 {shift['start_time'][:5]} 上班，實際 {r['first_in']}（晚 {late} 分鐘）"})
                except Exception:
                    pass
        if has_out and r['last_out'] and ds != str(today):
            shift = shift_map.get((r['staff_id'], ds))
            if shift and shift['end_time']:
                try:
                    eh, em = map(int, str(shift['end_time'])[:5].split(':'))
                    oh, om = map(int, r['last_out'].split(':'))
                    early = (eh*60+em)-(oh*60+om)
                    if early > 15:
                        anomalies.append({'type':'early','label':'早退','severity':'warning',
                                          'staff_id':r['staff_id'],'name':r['name'],'role':r['role'] or '',
                                          'department':r['department'] or '','date':ds,
                                          'detail':f"應 {shift['end_time'][:5]} 下班，實際 {r['last_out']}（早 {early} 分鐘）"})
                except Exception:
                    pass

    for s in all_staff:
        if s['id'] not in today_punched_ids and s['id'] not in on_leave_today_ids:
            anomalies.append({'type':'absent','label':'今日未出勤','severity':'error',
                               'staff_id':s['id'],'name':s['name'],'role':s['role'] or '',
                               'department':s['department'] or '','date':str(today),
                               'detail':'今日尚無打卡記錄且未請假'})

    sev_order = {'error':0,'warning':1,'info':2}
    anomalies.sort(key=lambda x: (sev_order.get(x['severity'],9), x['date']))
    result = {'anomalies':anomalies,'count':len(anomalies),'checked_from':str(date_from)}
    _anomalies_cache['data'] = result; _anomalies_cache['at'] = _an_now
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════
# Staff Termination
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/punch/staff/<int:sid>/terminate', methods=['POST'])
@login_required
def api_staff_terminate(sid):
    from routes.punch import punch_staff_row
    b = request.get_json(force=True)
    termination_date = b.get('termination_date','')
    reason = b.get('reason','').strip(); note = b.get('note','').strip()
    if not termination_date: return jsonify({'error': '請填寫離職日期'}), 400
    with get_db() as conn:
        try:
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_date DATE")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_reason TEXT DEFAULT ''")
            conn.execute("ALTER TABLE punch_staff ADD COLUMN IF NOT EXISTS termination_note TEXT DEFAULT ''")
        except Exception:
            pass
        row = conn.execute("""
            UPDATE punch_staff SET active=FALSE, termination_date=%s, termination_reason=%s,
              termination_note=%s, salary_notes=COALESCE(salary_notes,'')||%s
            WHERE id=%s RETURNING *
        """, (termination_date, reason, note, f'\n【離職】{termination_date} {reason}', sid)).fetchone()
        if not row: return ('', 404)
    return jsonify({'ok':True,'staff_id':sid,'name':row['name'],
                    'termination_date':termination_date,'last_salary_month':b.get('last_salary_month','')})


@bp.route('/api/punch/staff/<int:sid>/reinstate', methods=['POST'])
@login_required
def api_staff_reinstate(sid):
    from routes.punch import punch_staff_row
    with get_db() as conn:
        row = conn.execute("""
            UPDATE punch_staff SET active=TRUE,
              termination_date=NULL, termination_reason='', termination_note=''
            WHERE id=%s RETURNING *
        """, (sid,)).fetchone()
    return jsonify(punch_staff_row(row)) if row else ('', 404)


@bp.route('/api/punch/staff/terminated', methods=['GET'])
@login_required
def api_staff_terminated_list():
    with get_db() as conn:
        try:
            rows = conn.execute("""
                SELECT id, name, employee_code, department, role,
                       hire_date, termination_date, termination_reason
                FROM punch_staff WHERE active=FALSE
                ORDER BY termination_date DESC NULLS LAST, name
            """).fetchall()
        except Exception:
            rows = conn.execute(
                "SELECT id, name, employee_code, department, role, hire_date FROM punch_staff WHERE active=FALSE"
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for f in ('hire_date','termination_date'):
            if d.get(f): d[f] = str(d[f])
        result.append(d)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════
# Salary Formula Preview
# ═══════════════════════════════════════════════════════════════════

@bp.route('/api/salary/formula/preview', methods=['POST'])
@require_module('salary')
def api_formula_preview():
    from routes.salary import _eval_formula
    b              = request.get_json(force=True)
    formula        = b.get('formula','').strip()
    base_salary    = float(b.get('base_salary', 30000))
    insured_salary = float(b.get('insured_salary', 30000))
    service_years  = float(b.get('service_years', 1))
    if not formula: return jsonify({'result': 0, 'error': None})
    try:
        result = _eval_formula(formula, base_salary, insured_salary, service_years)
        return jsonify({'result': round(result, 2), 'error': None})
    except Exception as e:
        return jsonify({'result': None, 'error': str(e)})
