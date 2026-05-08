import json as _json
import time

from flask import Blueprint, Response, request, jsonify, session, render_template

from auth import login_required
from db import get_db, _qproducts_cache, _STATIC_TTL

bp = Blueprint('quotation', __name__)


def init():
    migrations = [
        """CREATE TABLE IF NOT EXISTS quotation_settings (
            id               SERIAL PRIMARY KEY,
            company_name     TEXT DEFAULT 'AD.Studio 影像事務所',
            company_address  TEXT DEFAULT '新北市三重區雙源街57巷7號1樓',
            company_phone    TEXT DEFAULT '(02)8985-1790',
            company_email    TEXT DEFAULT '',
            bank_name        TEXT DEFAULT '',
            bank_branch      TEXT DEFAULT '',
            account_name     TEXT DEFAULT '',
            account_no       TEXT DEFAULT '',
            tax_id           TEXT DEFAULT '',
            payment_notes    TEXT DEFAULT ''
        )""",
        """CREATE TABLE IF NOT EXISTS quotation_products (
            id            SERIAL PRIMARY KEY,
            name          TEXT NOT NULL,
            unit          TEXT DEFAULT '次',
            default_price NUMERIC(12,2) DEFAULT 0,
            sort_order    INTEGER DEFAULT 0,
            active        BOOLEAN DEFAULT TRUE
        )""",
        """CREATE TABLE IF NOT EXISTS quotations (
            id              SERIAL PRIMARY KEY,
            quote_no        TEXT UNIQUE NOT NULL,
            quote_date      DATE NOT NULL DEFAULT CURRENT_DATE,
            client_name     TEXT NOT NULL DEFAULT '',
            client_phone    TEXT DEFAULT '',
            client_address  TEXT DEFAULT '',
            client_line_id  TEXT DEFAULT '',
            sales_rep       TEXT DEFAULT '',
            image_no        TEXT DEFAULT '',
            image_date      DATE,
            payment_method  TEXT DEFAULT '轉帳',
            status          TEXT DEFAULT 'draft',
            subtotal        NUMERIC(12,2) DEFAULT 0,
            notes           TEXT DEFAULT '',
            created_by      TEXT DEFAULT '',
            created_at      TIMESTAMPTZ DEFAULT NOW(),
            updated_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS quotation_items (
            id             SERIAL PRIMARY KEY,
            quotation_id   INTEGER REFERENCES quotations(id) ON DELETE CASCADE,
            sort_order     INTEGER DEFAULT 0,
            product_name   TEXT NOT NULL DEFAULT '',
            unit           TEXT DEFAULT '次',
            quantity       NUMERIC(10,2) DEFAULT 1,
            unit_price     NUMERIC(12,2) DEFAULT 0,
            handmade       NUMERIC(12,2) DEFAULT 0,
            people_count   INTEGER DEFAULT 0,
            amount         NUMERIC(12,2) DEFAULT 0,
            payment_status TEXT DEFAULT '',
            note           TEXT DEFAULT ''
        )""",
        "ALTER TABLE quotation_settings ADD COLUMN IF NOT EXISTS company_unit TEXT DEFAULT 'ad'",
        "ALTER TABLE quotation_products ADD COLUMN IF NOT EXISTS company_unit TEXT DEFAULT 'ad'",
        "ALTER TABLE quotations         ADD COLUMN IF NOT EXISTS company_unit TEXT DEFAULT 'ad'",
        "ALTER TABLE quotations         ADD COLUMN IF NOT EXISTS deposit_rate NUMERIC DEFAULT 100",
        "ALTER TABLE quotations         ADD COLUMN IF NOT EXISTS show_wedding_content BOOLEAN DEFAULT TRUE",
        "ALTER TABLE quotation_settings ADD COLUMN IF NOT EXISTS frequent_accounts JSONB DEFAULT '[]'",
    ]
    for sql in migrations:
        try:
            with get_db() as conn:
                conn.execute(sql)
        except Exception as e:
            print(f"[quotation_init] {str(e)[:120]}")

    try:
        with get_db() as conn:
            for cu, cname, caddr, cphone in [
                ('ad', 'AD影像事務所', '新北市三重區雙源街57巷7號1樓', '(02)8985-1790'),
                ('jm', '進光設計', '', ''),
            ]:
                row = conn.execute(
                    "SELECT id FROM quotation_settings WHERE company_unit=%s LIMIT 1", (cu,)
                ).fetchone()
                if not row:
                    conn.execute(
                        "INSERT INTO quotation_settings (company_unit,company_name,company_address,company_phone) VALUES (%s,%s,%s,%s)",
                        (cu, cname, caddr, cphone)
                    )
    except Exception as e:
        print(f"[quotation_seed] {e}")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _next_quote_no(company_unit='ad'):
    from datetime import date as _d
    cu_prefix = 'JM' if company_unit == 'jm' else 'AD'
    prefix    = f'QT-{cu_prefix}-' + _d.today().strftime('%Y%m') + '-'
    with get_db() as conn:
        row = conn.execute(
            "SELECT quote_no FROM quotations WHERE quote_no LIKE %s ORDER BY quote_no DESC LIMIT 1",
            (prefix + '%',)
        ).fetchone()
    if row:
        try:
            seq = int(row['quote_no'].split('-')[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:03d}"


# ── Quote Number Preview ───────────────────────────────────────────────────────

@bp.route('/api/quotation/next-no', methods=['GET'])
@login_required
def api_quotation_next_no():
    company = request.args.get('company', 'ad')
    return jsonify({'quote_no': _next_quote_no(company)})


# ── Settings ──────────────────────────────────────────────────────────────────

@bp.route('/api/quotation/settings', methods=['GET'])
@login_required
def api_quotation_settings_get():
    company = request.args.get('company', 'ad')
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM quotation_settings WHERE company_unit=%s LIMIT 1", (company,)
        ).fetchone()
    return jsonify(dict(row) if row else {})


@bp.route('/api/quotation/settings', methods=['PUT'])
@login_required
def api_quotation_settings_put():
    b       = request.get_json(force=True) or {}
    company = b.get('company_unit', 'ad')
    fields  = ['company_name', 'company_address', 'company_phone', 'company_email',
               'bank_name', 'bank_branch', 'account_name', 'account_no', 'tax_id', 'payment_notes']
    fa = b.get('frequent_accounts', None)
    with get_db() as conn:
        row = conn.execute(
            "SELECT id FROM quotation_settings WHERE company_unit=%s LIMIT 1", (company,)
        ).fetchone()
        if row:
            sets = ', '.join(f"{f}=%s" for f in fields)
            vals = [b.get(f, '') for f in fields]
            if fa is not None:
                sets += ', frequent_accounts=%s'
                vals.append(_json.dumps(fa, ensure_ascii=False))
            vals.append(row['id'])
            conn.execute(f"UPDATE quotation_settings SET {sets} WHERE id=%s", vals)
        else:
            cols     = ', '.join(['company_unit'] + fields + (['frequent_accounts'] if fa is not None else []))
            phs      = ', '.join(['%s'] * (len(fields) + 1 + (1 if fa is not None else 0)))
            row_vals = [company] + [b.get(f, '') for f in fields]
            if fa is not None:
                row_vals.append(_json.dumps(fa, ensure_ascii=False))
            conn.execute(f"INSERT INTO quotation_settings ({cols}) VALUES ({phs})", row_vals)
    return jsonify({'ok': True})


# ── Preset Products ────────────────────────────────────────────────────────────

@bp.route('/api/quotation/products', methods=['GET'])
@login_required
def api_qproducts_list():
    company = request.args.get('company', '')
    key     = company or '__all__'
    now     = time.time()
    cached  = _qproducts_cache.get(key)
    if cached and now - cached['at'] < _STATIC_TTL:
        return jsonify(cached['data'])
    with get_db() as conn:
        if company:
            rows = conn.execute(
                "SELECT * FROM quotation_products WHERE company_unit=%s ORDER BY sort_order, id", (company,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM quotation_products ORDER BY sort_order, id"
            ).fetchall()
    result = [dict(r) for r in rows]
    _qproducts_cache[key] = {'data': result, 'at': now}
    return jsonify(result)


@bp.route('/api/quotation/products', methods=['POST'])
@login_required
def api_qproducts_create():
    b       = request.get_json(force=True) or {}
    company = b.get('company_unit', 'ad')
    with get_db() as conn:
        row = conn.execute(
            "INSERT INTO quotation_products (name, unit, default_price, sort_order, active, company_unit) VALUES (%s,%s,%s,%s,%s,%s) RETURNING *",
            (b.get('name', '').strip(), b.get('unit', '次'),
             float(b.get('default_price', 0)), int(b.get('sort_order', 0)), True, company)
        ).fetchone()
    _qproducts_cache.clear()
    return jsonify(dict(row)), 201


@bp.route('/api/quotation/products/<int:pid>', methods=['PUT'])
@login_required
def api_qproducts_update(pid):
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        conn.execute(
            "UPDATE quotation_products SET name=%s, unit=%s, default_price=%s, sort_order=%s, active=%s WHERE id=%s",
            (b.get('name', '').strip(), b.get('unit', '次'),
             float(b.get('default_price', 0)), int(b.get('sort_order', 0)),
             bool(b.get('active', True)), pid)
        )
    _qproducts_cache.clear()
    return jsonify({'ok': True})


@bp.route('/api/quotation/products/<int:pid>', methods=['DELETE'])
@login_required
def api_qproducts_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM quotation_products WHERE id=%s", (pid,))
    _qproducts_cache.clear()
    return jsonify({'ok': True})


# ── Quotations CRUD ────────────────────────────────────────────────────────────

@bp.route('/api/quotations', methods=['GET'])
@login_required
def api_quotations_list():
    status  = request.args.get('status', '')
    q       = request.args.get('q', '')
    company = request.args.get('company', '')
    conds, params = ['TRUE'], []
    if status:  conds.append("q.status=%s");                                      params.append(status)
    if q:       conds.append("(q.client_name ILIKE %s OR q.quote_no ILIKE %s)"); params += [f'%{q}%', f'%{q}%']
    if company: conds.append("q.company_unit=%s");                                params.append(company)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT q.*, COUNT(qi.id) as item_count
            FROM quotations q
            LEFT JOIN quotation_items qi ON qi.quotation_id = q.id
            WHERE {' AND '.join(conds)}
            GROUP BY q.id
            ORDER BY q.created_at DESC
        """, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in ('quote_date', 'image_date', 'created_at', 'updated_at'):
            if d.get(k): d[k] = str(d[k])
        result.append(d)
    return jsonify(result)


@bp.route('/api/quotations', methods=['POST'])
@login_required
def api_quotations_create():
    b          = request.get_json(force=True) or {}
    company    = b.get('company_unit', 'ad')
    quote_no   = _next_quote_no(company)
    from datetime import date as _dd
    quote_date = b.get('quote_date') or str(_dd.today())
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO quotations
              (quote_no, quote_date, client_name, client_phone, client_address,
               client_line_id, sales_rep, image_no, image_date, payment_method,
               status, notes, created_by, company_unit, deposit_rate, show_wedding_content)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (quote_no, quote_date,
              b.get('client_name', '').strip(),
              b.get('client_phone', '').strip(),
              b.get('client_address', '').strip(),
              b.get('client_line_id', '').strip(),
              b.get('sales_rep', '').strip(),
              b.get('image_no', '').strip(),
              b.get('image_date') or None,
              b.get('payment_method', '轉帳'),
              b.get('status', 'draft'),
              b.get('notes', '').strip(),
              session.get('admin_display_name', ''),
              company,
              float(b.get('deposit_rate', 100)),
              bool(b.get('show_wedding_content', True))
              )).fetchone()
        qid = row['id']
        for i, item in enumerate(b.get('items', [])):
            qty = float(item.get('quantity', 1))
            price = float(item.get('unit_price', 0))
            amt   = round(qty * price, 2)
            conn.execute("""
                INSERT INTO quotation_items
                  (quotation_id, sort_order, product_name, unit, quantity,
                   unit_price, handmade, people_count, amount, payment_status, note)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (qid, i,
                  item.get('product_name', '').strip(),
                  item.get('unit', '次'),
                  qty, price,
                  float(item.get('handmade', 0)),
                  int(item.get('people_count', 0)),
                  amt,
                  item.get('payment_status', ''),
                  item.get('note', '').strip()))
        conn.execute(
            "UPDATE quotations SET subtotal=(SELECT COALESCE(SUM(amount+handmade),0) FROM quotation_items WHERE quotation_id=%s) WHERE id=%s",
            (qid, qid)
        )
    return jsonify({'id': qid, 'quote_no': quote_no}), 201


@bp.route('/api/quotations/<int:qid>', methods=['GET'])
@login_required
def api_quotation_get(qid):
    with get_db() as conn:
        q = conn.execute("SELECT * FROM quotations WHERE id=%s", (qid,)).fetchone()
        if not q: return ('', 404)
        items = conn.execute(
            "SELECT * FROM quotation_items WHERE quotation_id=%s ORDER BY sort_order, id",
            (qid,)
        ).fetchall()
    d = dict(q)
    for k in ('quote_date', 'image_date', 'created_at', 'updated_at'):
        if d.get(k): d[k] = str(d[k])
    d['items'] = [dict(i) for i in items]
    return jsonify(d)


@bp.route('/api/quotations/<int:qid>', methods=['PUT'])
@login_required
def api_quotation_update(qid):
    b          = request.get_json(force=True) or {}
    new_status = b.get('status', 'draft')
    with get_db() as conn:
        old = conn.execute("SELECT status, subtotal, company_unit, quote_no, client_name FROM quotations WHERE id=%s", (qid,)).fetchone()
        conn.execute("""
            UPDATE quotations SET
              quote_date=%s, client_name=%s, client_phone=%s, client_address=%s,
              client_line_id=%s, sales_rep=%s, image_no=%s, image_date=%s,
              payment_method=%s, status=%s, notes=%s,
              deposit_rate=%s, show_wedding_content=%s, updated_at=NOW()
            WHERE id=%s
        """, (b.get('quote_date'), b.get('client_name', '').strip(),
              b.get('client_phone', '').strip(), b.get('client_address', '').strip(),
              b.get('client_line_id', '').strip(), b.get('sales_rep', '').strip(),
              b.get('image_no', '').strip(), b.get('image_date') or None,
              b.get('payment_method', '轉帳'), new_status,
              b.get('notes', '').strip(),
              float(b.get('deposit_rate', 100)),
              bool(b.get('show_wedding_content', True)),
              qid))
        # Auto-create finance income record when accepted
        if old and old['status'] != 'accepted' and new_status == 'accepted':
            subtotal = float(old['subtotal'] or 0)
            if subtotal > 0:
                company  = old['company_unit'] or 'ad'
                cat      = conn.execute(
                    "SELECT id FROM finance_categories WHERE type='income' AND company_unit=%s ORDER BY sort_order LIMIT 1",
                    (company,)
                ).fetchone()
                cat_id  = cat['id'] if cat else None
                already = conn.execute(
                    "SELECT id FROM finance_records WHERE linked_quotation_id=%s LIMIT 1", (qid,)
                ).fetchone()
                if not already:
                    from datetime import date as _dd2
                    conn.execute("""
                        INSERT INTO finance_records
                          (record_date, type, title, amount, category_id, company_unit, linked_quotation_id, note)
                        VALUES (%s,'income',%s,%s,%s,%s,%s,%s)
                    """, (str(_dd2.today()),
                          f"報價單收入 {old['quote_no']} - {old['client_name']}",
                          subtotal, cat_id, company, qid,
                          '由報價單自動建立'))
        conn.execute("DELETE FROM quotation_items WHERE quotation_id=%s", (qid,))
        for i, item in enumerate(b.get('items', [])):
            qty   = float(item.get('quantity', 1))
            price = float(item.get('unit_price', 0))
            amt   = round(qty * price, 2)
            conn.execute("""
                INSERT INTO quotation_items
                  (quotation_id, sort_order, product_name, unit, quantity,
                   unit_price, handmade, people_count, amount, payment_status, note)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (qid, i,
                  item.get('product_name', '').strip(),
                  item.get('unit', '次'),
                  qty, price,
                  float(item.get('handmade', 0)),
                  int(item.get('people_count', 0)),
                  amt,
                  item.get('payment_status', ''),
                  item.get('note', '').strip()))
        conn.execute(
            "UPDATE quotations SET subtotal=(SELECT COALESCE(SUM(amount+handmade),0) FROM quotation_items WHERE quotation_id=%s) WHERE id=%s",
            (qid, qid)
        )
    return jsonify({'ok': True})


@bp.route('/api/quotations/<int:qid>', methods=['DELETE'])
@login_required
def api_quotation_delete(qid):
    with get_db() as conn:
        conn.execute("DELETE FROM quotations WHERE id=%s", (qid,))
    return jsonify({'ok': True})


@bp.route('/api/quotations/<int:qid>/duplicate', methods=['POST'])
@login_required
def api_quotation_duplicate(qid):
    with get_db() as conn:
        q = conn.execute("SELECT * FROM quotations WHERE id=%s", (qid,)).fetchone()
        if not q: return ('', 404)
        items = conn.execute(
            "SELECT * FROM quotation_items WHERE quotation_id=%s ORDER BY sort_order", (qid,)
        ).fetchall()
        q      = dict(q)
        from datetime import date as _dd3
        new_no = _next_quote_no(q.get('company_unit', 'ad'))
        new_row = conn.execute("""
            INSERT INTO quotations
              (quote_no, quote_date, client_name, client_phone, client_address,
               client_line_id, sales_rep, image_no, image_date, payment_method,
               status, notes, created_by, company_unit, deposit_rate, show_wedding_content)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft',%s,%s,%s,%s,%s) RETURNING id
        """, (new_no, str(_dd3.today()),
              q.get('client_name', ''), q.get('client_phone', ''), q.get('client_address', ''),
              q.get('client_line_id', ''), q.get('sales_rep', ''),
              q.get('image_no', ''), q.get('image_date'),
              q.get('payment_method', '轉帳'),
              q.get('notes', ''), session.get('admin_display_name', ''),
              q.get('company_unit', 'ad'),
              float(q.get('deposit_rate') or 100),
              bool(q.get('show_wedding_content', True))
              )).fetchone()
        new_id = new_row['id']
        for item in items:
            item = dict(item)
            conn.execute("""
                INSERT INTO quotation_items
                  (quotation_id, sort_order, product_name, unit, quantity,
                   unit_price, handmade, people_count, amount, payment_status, note)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (new_id, item['sort_order'], item['product_name'], item['unit'],
                  item['quantity'], item['unit_price'], item['handmade'],
                  item['people_count'], item['amount'],
                  item['payment_status'], item['note']))
        conn.execute(
            "UPDATE quotations SET subtotal=(SELECT COALESCE(SUM(amount),0) FROM quotation_items WHERE quotation_id=%s) WHERE id=%s",
            (new_id, new_id)
        )
    return jsonify({'id': new_id, 'quote_no': new_no}), 201


# ── Clients CRUD ───────────────────────────────────────────────────────────────

@bp.route('/api/clients', methods=['GET'])
@login_required
def api_clients_list():
    q       = request.args.get('q', '')
    company = request.args.get('company', '')
    conds, params = ['TRUE'], []
    if company: conds.append("company_unit=%s"); params.append(company)
    if q:       conds.append("(name ILIKE %s OR phone ILIKE %s)"); params += [f'%{q}%', f'%{q}%']
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM clients WHERE {' AND '.join(conds)} ORDER BY name",
            params
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route('/api/clients', methods=['POST'])
@login_required
def api_clients_create():
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO clients (company_unit, name, phone, address, line_id, email, note)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
        """, (b.get('company_unit', 'ad'), b.get('name', '').strip(),
              b.get('phone', '').strip(), b.get('address', '').strip(),
              b.get('line_id', '').strip(), b.get('email', '').strip(),
              b.get('note', '').strip())).fetchone()
    return jsonify(dict(row)), 201


@bp.route('/api/clients/<int:cid>', methods=['PUT'])
@login_required
def api_clients_update(cid):
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        conn.execute("""
            UPDATE clients SET name=%s, phone=%s, address=%s, line_id=%s, email=%s, note=%s
            WHERE id=%s
        """, (b.get('name', '').strip(), b.get('phone', '').strip(),
              b.get('address', '').strip(), b.get('line_id', '').strip(),
              b.get('email', '').strip(), b.get('note', '').strip(), cid))
    return jsonify({'ok': True})


@bp.route('/api/clients/<int:cid>', methods=['DELETE'])
@login_required
def api_clients_delete(cid):
    with get_db() as conn:
        conn.execute("DELETE FROM clients WHERE id=%s", (cid,))
    return jsonify({'ok': True})


# ── Annual Revenue ─────────────────────────────────────────────────────────────

@bp.route('/api/finance/revenue-annual', methods=['GET'])
@login_required
def api_finance_revenue_annual():
    year    = request.args.get('year', '')
    company = request.args.get('company', '')
    if not year:
        from datetime import datetime as _dtm
        year = str(_dtm.now().year)
    c_cond  = "AND company_unit=%s" if company else ""
    c_param = [company] if company else []
    with get_db() as conn:
        this_rows = conn.execute(f"""
            SELECT to_char(record_date,'MM') as month, type,
                   COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE to_char(record_date,'YYYY')=%s {c_cond}
            GROUP BY 1,2
        """, [year] + c_param).fetchall()
        last_rows = conn.execute(f"""
            SELECT to_char(record_date,'MM') as month,
                   COALESCE(SUM(amount),0) as total
            FROM finance_records
            WHERE to_char(record_date,'YYYY')=%s AND type='income' {c_cond}
            GROUP BY 1
        """, [str(int(year) - 1)] + c_param).fetchall()
    inc  = {r['month']: float(r['total']) for r in this_rows if r['type'] == 'income'}
    exp  = {r['month']: float(r['total']) for r in this_rows if r['type'] == 'expense'}
    last = {r['month']: float(r['total']) for r in last_rows}
    months = [{'month': str(i).zfill(2),
               'income':      inc.get(str(i).zfill(2), 0),
               'expense':     exp.get(str(i).zfill(2), 0),
               'last_income': last.get(str(i).zfill(2), 0)} for i in range(1, 13)]
    return jsonify({'year': year, 'months': months})


# ── Print View ─────────────────────────────────────────────────────────────────

@bp.route('/quotation/<int:qid>/print')
@login_required
def quotation_print(qid):
    with get_db() as conn:
        q = conn.execute("SELECT * FROM quotations WHERE id=%s", (qid,)).fetchone()
        if not q: return '找不到此報價單', 404
        items        = conn.execute(
            "SELECT * FROM quotation_items WHERE quotation_id=%s ORDER BY sort_order, id",
            (qid,)
        ).fetchall()
        company_unit = dict(q).get('company_unit', 'ad')
        settings     = conn.execute(
            "SELECT * FROM quotation_settings WHERE company_unit=%s LIMIT 1", (company_unit,)
        ).fetchone()
        if not settings:
            settings = conn.execute("SELECT * FROM quotation_settings LIMIT 1").fetchone()
    qd = dict(q)
    for k in ('quote_date', 'image_date', 'created_at', 'updated_at'):
        if qd.get(k): qd[k] = str(qd[k])
    qd.setdefault('deposit_rate', 100)
    qd.setdefault('show_wedding_content', True)
    return render_template('quotation_print.html',
                           q=qd, items=[dict(i) for i in items],
                           s=dict(settings) if settings else {})


# ── Revenue Detail ─────────────────────────────────────────────────────────────

@bp.route('/api/revenue/detail', methods=['GET'])
@login_required
def api_revenue_detail():
    year          = request.args.get('year',   '')
    month         = request.args.get('month',  '')
    company       = request.args.get('company', '')
    status_filter = request.args.get('status', '')

    conds, params = ['TRUE'], []
    if year and not month:
        conds.append("to_char(q.quote_date,'YYYY')=%s"); params.append(year)
    if month:
        conds.append("to_char(q.quote_date,'YYYY-MM')=%s"); params.append(month)
    if company:
        conds.append("q.company_unit=%s"); params.append(company)
    if status_filter:
        statuses = status_filter.split(',')
        ph = ','.join(['%s'] * len(statuses))
        conds.append(f"q.status IN ({ph})"); params.extend(statuses)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                q.quote_no, q.quote_date, q.client_name, q.client_phone,
                q.image_no, q.image_date, q.payment_method,
                q.status, q.sales_rep, q.company_unit,
                qi.sort_order, qi.product_name, qi.unit, qi.quantity,
                qi.unit_price, qi.handmade, qi.people_count,
                qi.amount, qi.payment_status, qi.note
            FROM quotation_items qi
            JOIN quotations q ON q.id = qi.quotation_id
            WHERE {' AND '.join(conds)}
            ORDER BY q.quote_date DESC, q.id, qi.sort_order
        """, params).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        for k in ('quote_date', 'image_date'):
            if d.get(k): d[k] = str(d[k])
        for k in ('unit_price', 'handmade', 'amount', 'quantity'):
            if d.get(k) is not None: d[k] = float(d[k])
        result.append(d)
    return jsonify(result)


# ── Revenue Detail Excel Export ────────────────────────────────────────────────

@bp.route('/api/revenue/export', methods=['GET'])
@login_required
def api_revenue_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    year    = request.args.get('year',    '')
    month   = request.args.get('month',   '')
    company = request.args.get('company', '')

    conds, params = ['TRUE'], []
    if year and not month:
        conds.append("to_char(q.quote_date,'YYYY')=%s"); params.append(year)
    if month:
        conds.append("to_char(q.quote_date,'YYYY-MM')=%s"); params.append(month)
    if company:
        conds.append("q.company_unit=%s"); params.append(company)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                q.quote_no, q.quote_date, q.client_name,
                q.image_no, q.image_date, q.payment_method,
                q.status, q.sales_rep,
                qi.product_name, qi.unit, qi.quantity,
                qi.unit_price, qi.handmade, qi.people_count,
                qi.amount, qi.payment_status, qi.note
            FROM quotation_items qi
            JOIN quotations q ON q.id = qi.quotation_id
            WHERE {' AND '.join(conds)}
            ORDER BY q.quote_date DESC, q.id, qi.sort_order
        """, params).fetchall()

    status_map = {'draft': '草稿', 'sent': '已發送', 'accepted': '已接受', 'rejected': '已拒絕'}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '營業收入明細'

    hdr_fill = PatternFill('solid', fgColor='0F1C3A')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill = PatternFill('solid', fgColor='F4F6FA')
    center   = Alignment(horizontal='center', vertical='center')

    headers = ['訂單編號', '報價日期', '客戶', '影像編號', '影像日期',
               '訂單付款方式', '狀態', '業務人員',
               '品項目', '單位', '數量', '費率(元)', '手工費', '手操人數',
               '應收款項', '收款狀況', '備註']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24

    for i, r in enumerate(rows, 2):
        vals = [
            r['quote_no'], str(r['quote_date'] or ''),
            r['client_name'] or '', r['image_no'] or '',
            str(r['image_date'] or ''), r['payment_method'] or '',
            status_map.get(r['status'], r['status'] or ''), r['sales_rep'] or '',
            r['product_name'] or '', r['unit'] or '',
            float(r['quantity'] or 0), float(r['unit_price'] or 0),
            float(r['handmade'] or 0), int(r['people_count'] or 0),
            float(r['amount'] or 0),
            r['payment_status'] or '', r['note'] or '',
        ]
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            c.fill = fill
            c.alignment = Alignment(vertical='center')
            if col in (12, 13, 15):
                c.number_format = '#,##0'

    widths = [16, 12, 14, 14, 12, 12, 8, 10, 22, 6, 6, 10, 10, 8, 12, 12, 24]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    last  = len(rows) + 2
    total = sum(float(r['amount'] or 0) for r in rows)
    ws.cell(last, 1, f'共 {len(rows)} 筆').font = Font(bold=True)
    ws.cell(last, 15, total).font = Font(bold=True, color='2E9E6B')
    ws.cell(last, 15).number_format = '#,##0'

    buf   = BytesIO(); wb.save(buf); buf.seek(0)
    tag   = month or year or 'all'
    co_tag = f'_{company}' if company else ''
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=revenue{co_tag}_{tag}.xlsx'})


# ── Quotation Excel Export ─────────────────────────────────────────────────────

@bp.route('/api/quotation/export', methods=['GET'])
@login_required
def api_quotation_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    company = request.args.get('company', '')
    status  = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if company: conds.append("q.company_unit=%s"); params.append(company)
    if status:  conds.append("q.status=%s");        params.append(status)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT q.id, q.quote_no, q.quote_date, q.client_name, q.client_phone,
                   q.client_address, q.sales_rep, q.image_no, q.image_date,
                   q.payment_method, q.status, q.subtotal, q.notes,
                   q.company_unit, COUNT(qi.id) as item_count
            FROM quotations q
            LEFT JOIN quotation_items qi ON qi.quotation_id = q.id
            WHERE {' AND '.join(conds)}
            GROUP BY q.id
            ORDER BY q.created_at DESC
        """, params).fetchall()

        items_rows = conn.execute(f"""
            SELECT q.quote_no, q.client_name, qi.product_name, qi.unit, qi.quantity,
                   qi.unit_price, qi.handmade, qi.people_count, qi.amount,
                   qi.payment_status, qi.note
            FROM quotation_items qi
            JOIN quotations q ON q.id = qi.quotation_id
            WHERE {' AND '.join(conds)}
            ORDER BY q.created_at DESC, qi.sort_order
        """, params).fetchall()

    wb        = openpyxl.Workbook()
    hdr_fill  = PatternFill('solid', fgColor='0F1C3A')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    alt_fill  = PatternFill('solid', fgColor='F4F6FA')
    center    = Alignment(horizontal='center', vertical='center')
    status_map = {'draft': '草稿', 'sent': '已發送', 'accepted': '已接受', 'rejected': '已拒絕'}
    co_map     = {'ad': 'AD影像事務所', 'jm': '進光設計'}

    ws1 = wb.active
    ws1.title = '報價單列表'
    hdrs1 = ['訂單編號', '報價日期', '客戶', '聯絡電話', '業務人員', '影像編號',
             '付款方式', '狀態', '金額', '品項數', '公司單位']
    for col, h in enumerate(hdrs1, 1):
        c = ws1.cell(1, col, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws1.freeze_panes = 'A2'
    ws1.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 2):
        vals = [
            r['quote_no'], str(r['quote_date'] or ''),
            r['client_name'] or '', r['client_phone'] or '',
            r['sales_rep'] or '', r['image_no'] or '',
            r['payment_method'] or '',
            status_map.get(r['status'], r['status'] or ''),
            float(r['subtotal'] or 0), int(r['item_count'] or 0),
            co_map.get(r['company_unit'] or '', r['company_unit'] or ''),
        ]
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col, v in enumerate(vals, 1):
            c = ws1.cell(i, col, v)
            c.fill = fill
            c.alignment = Alignment(vertical='center')
            if col == 9:
                c.number_format = '#,##0'
                c.font = Font(bold=True)

    for col, w in enumerate([18, 12, 16, 14, 12, 16, 10, 10, 14, 8, 14], 1):
        ws1.column_dimensions[ws1.cell(1, col).column_letter].width = w

    last1     = len(rows) + 2
    total_amt = sum(float(r['subtotal'] or 0) for r in rows)
    ws1.cell(last1, 1, f'共 {len(rows)} 筆').font = Font(bold=True)
    ws1.cell(last1, 9, total_amt).font = Font(bold=True, color='2E9E6B')
    ws1.cell(last1, 9).number_format = '#,##0'

    ws2 = wb.create_sheet('品項明細')
    hdrs2 = ['訂單編號', '客戶', '產品名稱', '單位', '數量', '費率(元)', '手工費', '人數', '費用', '收款狀況', '備註']
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(1, col, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 22

    for i, r in enumerate(items_rows, 2):
        vals2 = [
            r['quote_no'], r['client_name'] or '',
            r['product_name'] or '', r['unit'] or '',
            float(r['quantity'] or 0), float(r['unit_price'] or 0),
            float(r['handmade'] or 0), int(r['people_count'] or 0),
            float(r['amount'] or 0),
            r['payment_status'] or '', r['note'] or '',
        ]
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col, v in enumerate(vals2, 1):
            c = ws2.cell(i, col, v)
            c.fill = fill
            c.alignment = Alignment(vertical='center')
            if col in (5, 6, 7, 9): c.number_format = '#,##0.##'

    for col, w in enumerate([18, 16, 24, 8, 8, 12, 10, 8, 12, 12, 20], 1):
        ws2.column_dimensions[ws2.cell(1, col).column_letter].width = w

    buf    = BytesIO(); wb.save(buf); buf.seek(0)
    co_tag = f'_{company}' if company else ''
    fname  = f"quotations{co_tag}.xlsx"
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})
